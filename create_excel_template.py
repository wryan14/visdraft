import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def create_excel_template(filename):
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    required_font = Font(color="FF0000", bold=True)
    instruction_font = Font(color="666666", italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Create Main Info sheet
    ws_main = wb.active
    ws_main.title = "Basic Information"
    
    # Headers and fields for Basic Information
    basic_info_fields = [
        ("Dataset Name", "The unique identifier for this dataset", True),
        ("Owner/Maintainer", "Primary person responsible for this dataset", True),
        ("Last Updated", "Date of last update (YYYY-MM-DD)", True),
        ("Update Frequency", "How often is the dataset updated", True),
        ("Business Purpose", "Primary business use of this dataset", True),
        ("Dataset Overview", "Detailed description of what this dataset represents", True),
        ("Source System", "Origin system of the data", True),
        ("Data Collection Method", "How is the data collected/generated", True),
        ("Source Data Format", "Original format of the data", True),
    ]

    ws_main.column_dimensions['A'].width = 25
    ws_main.column_dimensions['B'].width = 50
    ws_main.column_dimensions['C'].width = 40

    # Add headers
    ws_main['A1'] = "Field"
    ws_main['B1'] = "Value"
    ws_main['C1'] = "Description"
    
    for cell in ws_main[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Add fields
    for idx, (field, description, required) in enumerate(basic_info_fields, start=2):
        ws_main[f'A{idx}'] = field
        ws_main[f'C{idx}'] = description
        if required:
            ws_main[f'A{idx}'].font = required_font
    
    # Create Schema Details sheet
    ws_schema = wb.create_sheet("Schema Details")
    
    schema_headers = ["Column Name", "Data Type", "Description", "Constraints/Rules", "Example Value"]
    schema_descriptions = [
        "Name of the column in the dataset",
        "Data type (string, integer, date, etc.)",
        "What this column represents",
        "Any rules or constraints on the data",
        "Representative example of the data"
    ]
    
    for col, (header, desc) in enumerate(zip(schema_headers, schema_descriptions), start=1):
        col_letter = get_column_letter(col)
        ws_schema[f'{col_letter}1'] = header
        ws_schema[f'{col_letter}2'] = desc
        ws_schema.column_dimensions[col_letter].width = 25
    
    # Style the schema headers
    for cell in ws_schema[1]:
        cell.fill = header_fill
        cell.font = header_font
    
    # Style the descriptions
    for cell in ws_schema[2]:
        cell.font = instruction_font
        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

    # Create Data Quality sheet
    ws_quality = wb.create_sheet("Data Quality")
    
    quality_sections = [
        ("Validation Rules", "List all business rules and validations applied to the dataset"),
        ("Known Issues", "Document any known data quality issues or limitations"),
        ("Access Method", "How to access this dataset"),
        ("Authentication Required", "Yes/No"),
        ("Authorization Level", "Required access level"),
        ("Sensitive Data Categories", "PII, PHI, etc. if any"),
    ]
    
    for idx, (section, description) in enumerate(quality_sections, start=1):
        ws_quality[f'A{idx*2-1}'] = section
        ws_quality[f'A{idx*2}'] = description
        ws_quality.row_dimensions[idx*2-1].height = 30
        ws_quality.row_dimensions[idx*2].height = 50
        
        # Style headers
        ws_quality[f'A{idx*2-1}'].fill = header_fill
        ws_quality[f'A{idx*2-1}'].font = header_font
        ws_quality[f'A{idx*2}'].font = instruction_font
    
    ws_quality.column_dimensions['A'].width = 30
    ws_quality.column_dimensions['B'].width = 60

    # Create Dependencies sheet
    ws_deps = wb.create_sheet("Dependencies")
    
    deps_sections = [
        ("Upstream Dependencies", "List systems/processes this dataset depends on"),
        ("Downstream Dependencies", "List systems/processes that depend on this dataset"),
        ("Processing Schedule", "When and how often is the data processed"),
        ("Critical Timings", "Any time-sensitive dependencies"),
    ]
    
    for idx, (section, description) in enumerate(deps_sections, start=1):
        ws_deps[f'A{idx*2-1}'] = section
        ws_deps[f'A{idx*2}'] = description
        ws_deps.row_dimensions[idx*2-1].height = 30
        ws_deps.row_dimensions[idx*2].height = 50
        
        ws_deps[f'A{idx*2-1}'].fill = header_fill
        ws_deps[f'A{idx*2-1}'].font = header_font
        ws_deps[f'A{idx*2}'].font = instruction_font
    
    ws_deps.column_dimensions['A'].width = 30
    ws_deps.column_dimensions['B'].width = 60

    # Create Change Log sheet
    ws_changes = wb.create_sheet("Change Log")
    
    change_headers = ["Date", "Change Description", "Changed By"]
    for col, header in enumerate(change_headers, start=1):
        ws_changes[f'{get_column_letter(col)}1'] = header
        ws_changes.column_dimensions[get_column_letter(col)].width = 25
        ws_changes[f'{get_column_letter(col)}1'].fill = header_fill
        ws_changes[f'{get_column_letter(col)}1'].font = header_font

    # Save the workbook
    wb.save(filename)

# Create the template
create_excel_template('/tmp/dataset_schema_template.xlsx')